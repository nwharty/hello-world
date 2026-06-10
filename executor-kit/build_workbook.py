#!/usr/bin/env python3
"""Generates Executor_Estate_Kit.xlsx — a 12-tab estate administration workbook.

Visual language follows the Executor Kit design system (design_handoff):
warm paper neutrals, navy headers, sage chips, brass accents, zebra rows.
The same .xlsx uploads cleanly to Google Sheets (File > Import).
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --- design tokens (brand.css) ---
NAVY = "223A52"
NAVY_DEEP = "182B3E"
PAPER = "FAF7F1"
ZEBRA = "F9F4EB"        # paper-2 at ~45% over white
WHITE = "FFFDF9"
SAGE = "E7EFE5"
SAGE_DEEP = "C4D6C0"
SAGE_INK = "4A6450"
GOLD = "B58A34"
GOLD_PALE = "ECDCB4"
GOLD_TEXT = "7C5D1A"
CLAY = "BD7B5C"
OK = "5D8A63"
RED = "B00020"
INK_55 = "6B7E91"        # navy at ~55% on paper, flattened
CHIP_TODO = "EEF1F4"

SERIF = "Georgia"        # Spectral fallback per handoff
SANS = "Calibri"         # Hanken Grotesk fallback
MONO = "Consolas"        # Spline Sans Mono fallback

HEADER_FONT = Font(name=SANS, bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(name=SERIF, bold=False, size=18, color=NAVY)
SUB_FONT = Font(name=SANS, size=11, color=INK_55)
EYEBROW_FONT = Font(name=MONO, bold=True, size=9, color=GOLD)
SECTION_FONT = Font(name=SERIF, bold=True, size=12, color=NAVY)
PHASE_FONT = Font(name=MONO, size=9, color=GOLD)
THIN = Side(style="thin", color="DDD6C9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NAVY_TOP = Border(left=THIN, right=THIN, top=Side(style="medium", color=NAVY), bottom=THIN)
MONEY = '"$"#,##0.00'

DISCLAIMER = (
    "This workbook is an organizational tool, not legal, tax, or financial advice. "
    "Estate and probate rules vary by state and country - confirm requirements with "
    "the probate court or a licensed professional."
)


def sheet_title(ws, title, subtitle, ncols, eyebrow=None):
    r = 1
    if eyebrow:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(row=1, column=1, value=eyebrow.upper()).font = EYEBROW_FONT
        r = 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    ws.cell(row=r, column=1, value=title).font = TITLE_FONT
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=ncols)
    ws.cell(row=r + 1, column=1, value=subtitle).font = SUB_FONT
    ws.row_dimensions[r].height = 30
    ws.sheet_view.showGridLines = False


def header_row(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def blank_rows(ws, start, count, ncols, money_cols=(), date_cols=()):
    for r in range(start, start + count):
        for col in range(1, ncols + 1):
            c = ws.cell(row=r, column=col)
            c.border = BORDER
            c.font = Font(name=SANS, size=11, color=NAVY)
            if col in money_cols:
                c.number_format = MONEY
            if col in date_cols:
                c.number_format = "mm/dd/yyyy"
        if (r - start) % 2 == 1:
            for col in range(1, ncols + 1):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=ZEBRA)


def dropdown(ws, options, col_letter, first, last):
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first}:{col_letter}{last}")


def total_cell(ws, row, col, formula):
    c = ws.cell(row=row, column=col, value=formula)
    c.number_format = MONEY
    c.font = Font(name=SANS, bold=True, size=11, color=NAVY)
    c.fill = PatternFill("solid", fgColor=SAGE)
    c.border = NAVY_TOP
    return c


STATUS = ["Not started", "In progress", "Done", "N/A"]

wb = Workbook()

# ---------------------------------------------------------------- Start Here
ws = wb.active
ws.title = "Start Here"
ws.sheet_properties.tabColor = GOLD
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 5
ws.column_dimensions["C"].width = 72
ws.column_dimensions["D"].width = 4
ws.column_dimensions["E"].width = 46

ws.cell(row=2, column=2, value="✦").font = Font(name=SERIF, size=22, color=GOLD)
ws.cell(row=3, column=2)
ws.merge_cells("B4:C4")
ws.cell(row=4, column=2, value="Executor & Estate Settlement Kit").font = Font(name=SERIF, size=22, color=NAVY)
ws.merge_cells("B5:C5")
ws.cell(row=5, column=2, value="A calm, complete system for settling an estate - one tab at a time.").font = SUB_FONT
ws.row_dimensions[4].height = 30

ws.cell(row=7, column=2, value="HOW TO USE THIS KIT").font = EYEBROW_FONT
steps = [
    "Read the companion guide first - it tells you what's urgent and what can wait.",
    "Work the First 30 Days tab, top to bottom. You don't have to do everything at once.",
    "Log every call, email, and letter. Months from now, this record protects you.",
    "Record each asset, debt, and document as you find them. Values can come later.",
    "Track every dollar in the Estate Ledger. Final Accounting totals it for you.",
]
r = 8
for i, s in enumerate(steps, 1):
    n = ws.cell(row=r, column=2, value=i)
    n.font = Font(name=SERIF, bold=True, size=11, color=SAGE_INK)
    n.fill = PatternFill("solid", fgColor=SAGE)
    n.alignment = Alignment(horizontal="center", vertical="center")
    c = ws.cell(row=r, column=3, value=s)
    c.font = Font(name=SANS, size=11, color=NAVY)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[r].height = 28
    r += 1

# right column: tab directory
ws.cell(row=7, column=5, value="THE TABS").font = EYEBROW_FONT
tab_dir = [
    ("First 30 Days", NAVY, "What's urgent, what can wait"),
    ("Task Log", None, "Every call & letter, documented"),
    ("Key Contacts", None, "Everyone you'll call twice"),
    ("Documents", None, "Where every paper lives"),
    ("Notifications", None, "Who to tell + cert. copies"),
    ("Assets", GOLD, "Full date-of-death inventory"),
    ("Debts", None, "Verified before a dollar is paid"),
    ("Services", None, "Subscriptions to cancel"),
    ("Estate Ledger", GOLD, "Every dollar in and out"),
    ("Distributions", None, "Who received what, signed"),
    ("Final Accounting", RED, "The court summary, automatic"),
]
r2 = 8
for name, color, desc in tab_dir:
    c = ws.cell(row=r2, column=5, value=f"{name}  -  {desc}")
    c.font = Font(name=SANS, size=10.5, bold=bool(color), color=color or INK_55)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    r2 += 1

gs = ws.cell(row=r2 + 1, column=5, value="GOOGLE SHEETS - upload this file at sheets.google.com via File > Import. Every formula works.")
gs.font = Font(name=MONO, size=9, color=SAGE_INK)
gs.fill = PatternFill("solid", fgColor=SAGE)
gs.alignment = Alignment(wrap_text=True, vertical="center")
ws.row_dimensions[r2 + 1].height = 44

d = ws.cell(row=max(r, r2) + 3, column=2, value=DISCLAIMER)
ws.merge_cells(start_row=max(r, r2) + 3, start_column=2, end_row=max(r, r2) + 3, end_column=3)
d.font = Font(name=SANS, size=9, italic=True, color=INK_55)
d.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[max(r, r2) + 3].height = 40

# ------------------------------------------------------------- First 30 Days
ws = wb.create_sheet("First 30 Days")
ws.sheet_properties.tabColor = NAVY
sheet_title(ws, "The First 30 Days", "Work top to bottom. 'Done' and 'N/A' both count as finished.", 5, eyebrow="Know what's urgent - and what can wait")
header_row(ws, 5, ["Priority", "Task", "Why it matters", "Status", "Notes"], [11, 52, 52, 14, 30])

tasks = [
    ("WEEK 1", "Locate the original will (and any codicils)", "Check safe, filing cabinet, desk, attorney, bank safe-deposit box. The original is usually required by the court."),
    ("WEEK 1", "Order 10-15 certified copies of the death certificate", "Banks, insurers, and agencies each demand their own certified copy. Running out mid-process causes weeks of delay."),
    ("WEEK 1", "Secure the home, vehicles, and valuables", "You are responsible for protecting estate property from this point forward. Change locks if needed."),
    ("WEEK 1", "Arrange care for dependents and pets", "Immediate welfare comes before paperwork."),
    ("WEEK 1", "Forward mail (USPS) and collect incoming bills", "Mail reveals accounts, debts, and subscriptions you do not know about yet."),
    ("WEEK 1", "Do NOT pay estate bills from your own money, and do NOT distribute anything yet", "Commingling funds and early distributions are the two most common - and most personally costly - executor mistakes."),
    ("WEEK 2", "File the will and petition for probate with the county court", "Nothing official can happen until the court appoints you. Ask the clerk what your county requires."),
    ("WEEK 2", "Obtain Letters Testamentary / Letters of Administration", "This court document is your legal authority. Institutions will refuse to talk to you without it."),
    ("WEEK 2", "Get an EIN for the estate (irs.gov, free, ~10 minutes)", "The estate is its own taxpayer. You need an EIN before opening the estate bank account."),
    ("WEEK 2", "Open an estate checking account", "Every dollar in and out of the estate flows through this one account. Start the Estate Ledger tab the same day."),
    ("WEEK 2-3", "Notify Social Security, employer, insurers, and banks", "Use the Notifications tab. Benefits may need to be stopped or claimed; accounts must be frozen or retitled."),
    ("WEEK 2-3", "Notify the three credit bureaus and request a credit report", "Prevents identity theft and reveals unknown debts and accounts."),
    ("WEEK 3-4", "Begin the asset inventory with date-of-death values", "The court's inventory filing and the final accounting are both built on these numbers. Use the Assets tab."),
    ("WEEK 3-4", "Identify and verify all debts before paying anything", "Debts are paid in a legal order of priority. Paying the wrong ones first can make you personally liable."),
    ("WEEK 3-4", "Cancel or transfer subscriptions and services", "Use the Services tab. Every month of delay drains the estate."),
    ("WEEK 3-4", "Calendar the tax deadlines (final 1040, estate 1041)", "The final personal return and the estate income tax return have firm deadlines with penalties."),
    ("ONGOING", "Log every action in the Task Log; keep every receipt", "Beneficiaries are entitled to an accounting. A complete record is your best protection against disputes."),
    ("ONGOING", "Consider a probate attorney or CPA for anything unclear", "Reasonable professional fees are paid by the estate, not by you. Getting help is normal, not failure."),
]
r = 6
for phase, task, why in tasks:
    p = ws.cell(row=r, column=1, value=phase)
    p.font = PHASE_FONT
    p.border = BORDER
    p.alignment = Alignment(vertical="center")
    c = ws.cell(row=r, column=2, value=task)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.font = Font(name=SANS, size=11, bold=True, color=NAVY)
    c.border = BORDER
    c = ws.cell(row=r, column=3, value=why)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    c.font = Font(name=SANS, size=10, color=INK_55)
    c.border = BORDER
    ws.cell(row=r, column=4).border = BORDER
    ws.cell(row=r, column=5).border = BORDER
    if (r - 6) % 2 == 1:
        for col in range(1, 6):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=ZEBRA)
    ws.row_dimensions[r].height = 42
    r += 1
dropdown(ws, STATUS, "D", 6, r - 1)

# ------------------------------------------------------------------ Task Log
ws = wb.create_sheet("Task Log")
sheet_title(ws, "Task & Communication Log", "One row per call, email, or letter. This log is your protection.", 7, eyebrow="Write everything down")
header_row(ws, 5, ["Date", "Type", "Who / Organization", "Regarding", "Outcome", "Follow-up date", "Follow-up done?"], [13, 12, 26, 34, 34, 15, 15])
blank_rows(ws, 6, 120, 7, date_cols=(1, 6))
dropdown(ws, ["Call", "Email", "Letter", "In person", "Court filing", "Other"], "B", 6, 125)
dropdown(ws, ["Yes", "No", "N/A"], "G", 6, 125)

# -------------------------------------------------------------- Key Contacts
ws = wb.create_sheet("Key Contacts")
sheet_title(ws, "Key Contacts", "Everyone you will call more than once.", 6, eyebrow="Keep them one tab away")
header_row(ws, 5, ["Role", "Name", "Organization", "Phone", "Email", "Notes / account or case #"], [24, 22, 26, 17, 28, 34])
roles = ["Probate attorney", "CPA / tax preparer", "Probate court clerk", "Funeral home", "Financial advisor",
         "Life insurance agent", "Realtor / appraiser", "Employer HR contact", "Beneficiary", "Beneficiary", "Beneficiary"]
r = 6
for role in roles:
    c = ws.cell(row=r, column=1, value=role)
    c.font = Font(name=SANS, size=11, bold=True, color=NAVY)
    c.border = BORDER
    for col in range(2, 7):
        ws.cell(row=r, column=col).border = BORDER
    if (r - 6) % 2 == 1:
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=ZEBRA)
    r += 1
blank_rows(ws, r, 25, 6)

# ----------------------------------------------------------------- Documents
ws = wb.create_sheet("Documents")
sheet_title(ws, "Document Locator", "Where every important paper lives. Fill in locations as you find them.", 5, eyebrow="Where every paper lives")
header_row(ws, 5, ["Document", "Located?", "Where it is / where found", "Original or copy?", "Notes"], [38, 12, 36, 16, 30])
docs = ["Original will / codicils", "Trust documents", "Death certificates (certified)", "Letters Testamentary",
        "Birth certificate", "Marriage certificate", "Social Security card", "Deeds to real estate",
        "Vehicle titles", "Bank statements", "Investment / brokerage statements", "Retirement account statements",
        "Life insurance policies", "Pension / annuity documents", "Last 3 years of tax returns",
        "Mortgage / loan documents", "Business ownership documents", "Safe-deposit box key & location",
        "Password list / digital accounts", "Prepaid funeral contract", "Military discharge (DD-214)"]
r = 6
for d_ in docs:
    c = ws.cell(row=r, column=1, value=d_)
    c.font = Font(name=SANS, size=11, bold=True, color=NAVY)
    c.border = BORDER
    for col in range(2, 6):
        ws.cell(row=r, column=col).border = BORDER
    if (r - 6) % 2 == 1:
        for col in range(1, 6):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=ZEBRA)
    r += 1
blank_rows(ws, r, 10, 5)
dropdown(ws, ["Yes", "No", "Searching"], "B", 6, r + 9)
dropdown(ws, ["Original", "Copy"], "D", 6, r + 9)

# ------------------------------------------------------------- Notifications
ws = wb.create_sheet("Notifications")
sheet_title(ws, "Notifications & Death Certificate Tracker", "Who must be told - and which ones need a certified death certificate.", 7, eyebrow="The one thing every executor runs out of")
header_row(ws, 5, ["Who to notify", "Certified copy required?", "Date notified", "Method", "Confirmed / reference #", "Cert. copies used", "Notes"], [34, 20, 14, 14, 24, 16, 28])
parties = ["Social Security Administration", "Employer / HR (final pay, benefits)", "Life insurance company",
           "Health insurance / Medicare", "Banks (each account)", "Brokerage / investment firms",
           "Retirement plan administrators", "Pension provider", "Mortgage company", "Credit card issuers",
           "Credit bureaus (Equifax, Experian, TransUnion)", "DMV (licenses, vehicle titles)", "Veterans Affairs (if applicable)",
           "Utility companies", "Landlord / HOA", "Post office (mail forwarding)"]
r = 6
for p_ in parties:
    c = ws.cell(row=r, column=1, value=p_)
    c.font = Font(name=SANS, size=11, bold=True, color=NAVY)
    c.border = BORDER
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=3).number_format = "mm/dd/yyyy"
    if (r - 6) % 2 == 1:
        for col in range(1, 8):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=ZEBRA)
    r += 1
blank_rows(ws, r, 10, 7, date_cols=(3,))
last = r + 9
dropdown(ws, ["Yes", "No", "Unsure"], "B", 6, last)
dropdown(ws, ["Call", "Mail", "Online", "In person"], "D", 6, last)
lbl = ws.cell(row=last + 2, column=5, value="Certified copies used so far:")
lbl.font = Font(name=SANS, bold=True, size=11, color=NAVY)
lbl.alignment = Alignment(horizontal="right")
total_cell(ws, last + 2, 6, f"=SUM(F6:F{last})").number_format = "0"

# -------------------------------------------------------------------- Assets
ws = wb.create_sheet("Assets")
ws.sheet_properties.tabColor = GOLD
sheet_title(ws, "Asset Inventory", "Every asset, valued as of the date of death - the foundation of the inventory filing and final accounting.", 9, eyebrow="The beginning inventory")
header_row(ws, 5, ["Category", "Institution / Description", "Account # (last 4)", "How titled",
                   "Has named beneficiary?", "Value at date of death", "Current value", "Status", "Notes"],
           [22, 30, 14, 18, 18, 18, 18, 16, 28])
blank_rows(ws, 6, 60, 9, money_cols=(6, 7))
cats = ["Bank account", "Investment / brokerage", "Retirement (IRA/401k)", "Real estate", "Vehicle",
        "Life insurance", "Business interest", "Personal property", "Digital asset", "Other"]
dropdown(ws, cats, "A", 6, 65)
dropdown(ws, ["Sole name", "Joint", "In trust", "POD/TOD"], "D", 6, 65)
dropdown(ws, ["Yes", "No", "Unsure"], "E", 6, 65)
dropdown(ws, ["Located", "Valued", "Retitled", "Sold", "Distributed", "Closed"], "H", 6, 65)
t = ws.cell(row=67, column=5, value="TOTAL - beginning inventory:")
t.font = Font(name=SANS, bold=True, size=11, color=NAVY)
t.fill = PatternFill("solid", fgColor=SAGE)
t.border = NAVY_TOP
t.alignment = Alignment(horizontal="right")
for col in (6, 7):
    total_cell(ws, 67, col, f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}65)")

# --------------------------------------------------------------------- Debts
ws = wb.create_sheet("Debts")
sheet_title(ws, "Debts & Liabilities", "Verify every claim before paying. Debts are paid in a legal order of priority - ask the court or an attorney if assets may not cover them all.", 8, eyebrow="Verified before a dollar is paid")
header_row(ws, 5, ["Creditor", "Type", "Account # (last 4)", "Balance at death", "Verified?", "Amount paid", "Date paid", "Notes"],
           [28, 20, 14, 16, 12, 16, 13, 30])
blank_rows(ws, 6, 40, 8, money_cols=(4, 6), date_cols=(7,))
dropdown(ws, ["Mortgage", "Auto loan", "Credit card", "Medical", "Taxes", "Personal loan", "Utility", "Funeral", "Other"], "B", 6, 45)
dropdown(ws, ["Yes", "No", "Disputed"], "E", 6, 45)
t = ws.cell(row=47, column=3, value="TOTALS:")
t.font = Font(name=SANS, bold=True, size=11, color=NAVY)
t.fill = PatternFill("solid", fgColor=SAGE)
t.border = NAVY_TOP
t.alignment = Alignment(horizontal="right")
for col in (4, 6):
    total_cell(ws, 47, col, f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}45)")

# ------------------------------------------------------------------ Services
ws = wb.create_sheet("Services")
sheet_title(ws, "Services & Subscriptions", "Every recurring charge to cancel or transfer. Check bank and card statements for the full list.", 7, eyebrow="Stop the slow leaks")
header_row(ws, 5, ["Service", "Account / login", "Monthly cost", "Action", "Date completed", "Refund due?", "Notes"],
           [28, 26, 14, 16, 15, 13, 30])
svcs = ["Electric / gas", "Water / sewer", "Internet / cable", "Cell phone", "Streaming services",
        "Newspaper / magazines", "Gym membership", "Insurance (auto/home)", "Amazon / shopping", "Cloud storage"]
r = 6
for s in svcs:
    c = ws.cell(row=r, column=1, value=s)
    c.font = Font(name=SANS, size=11, bold=True, color=NAVY)
    c.border = BORDER
    for col in range(2, 8):
        ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=3).number_format = MONEY
    ws.cell(row=r, column=5).number_format = "mm/dd/yyyy"
    if (r - 6) % 2 == 1:
        for col in range(1, 8):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=ZEBRA)
    r += 1
blank_rows(ws, r, 15, 7, money_cols=(3,), date_cols=(5,))
dropdown(ws, ["Cancel", "Transfer", "Keep (estate)", "Done"], "D", 6, r + 14)

# ------------------------------------------------------------- Estate Ledger
ws = wb.create_sheet("Estate Ledger")
ws.sheet_properties.tabColor = GOLD
sheet_title(ws, "Estate Ledger - Money In / Money Out", "Every transaction in the estate account. Final Accounting totals this automatically.", 7, eyebrow="Every dollar, accounted for")
header_row(ws, 5, ["Date", "Type", "Category", "Payer / Payee", "Amount", "Receipt kept?", "Notes"],
           [13, 16, 24, 28, 16, 13, 34])
blank_rows(ws, 6, 150, 7, money_cols=(5,), date_cols=(1,))
dropdown(ws, ["Receipt (money in)", "Disbursement (money out)"], "B", 6, 155)
dropdown(ws, ["Asset sale proceeds", "Interest / dividends", "Refunds", "Final paycheck", "Other income",
              "Funeral expense", "Court / filing fees", "Attorney / CPA fees", "Taxes paid", "Debt payment",
              "Property upkeep", "Executor expense", "Other expense"], "C", 6, 155)
dropdown(ws, ["Yes", "No"], "F", 6, 155)

# ------------------------------------------------------------- Distributions
ws = wb.create_sheet("Distributions")
sheet_title(ws, "Beneficiary Distributions", "Distribute only after debts and taxes are settled (or with court/attorney guidance). Get a signed receipt for everything.", 8, eyebrow="Who received what, signed")
header_row(ws, 5, ["Beneficiary", "Relationship", "Entitlement per will", "Asset / item distributed", "Value / amount", "Date", "Receipt or release signed?", "Notes"],
           [24, 16, 24, 28, 16, 13, 12, 28])
blank_rows(ws, 6, 40, 8, money_cols=(5,), date_cols=(6,))
dropdown(ws, ["Yes", "No", "Sent - awaiting"], "G", 6, 45)
t = ws.cell(row=47, column=4, value="TOTAL DISTRIBUTED:")
t.font = Font(name=SANS, bold=True, size=11, color=NAVY)
t.fill = PatternFill("solid", fgColor=SAGE)
t.border = NAVY_TOP
t.alignment = Alignment(horizontal="right")
total_cell(ws, 47, 5, "=SUM(E6:E45)")

# ---------------------------------------------------------- Final Accounting
ws = wb.create_sheet("Final Accounting")
ws.sheet_properties.tabColor = RED
sheet_title(ws, "Final Accounting Summary",
            "Calculated automatically from your other tabs - the structure courts and beneficiaries expect: what came in, what went out, what remains.",
            3, eyebrow="The hardest report writes itself")
ws.column_dimensions["A"].width = 52
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 60

lines = [
    ("ESTATE AT A GLANCE", None, ""),
    ("Beginning inventory (assets at date-of-death value)", "=SUM(Assets!F6:F65)", "From the Assets tab"),
    ("Plus: receipts during administration", "=SUMIF('Estate Ledger'!B6:B155,\"Receipt (money in)\",'Estate Ledger'!E6:E155)", "From the Estate Ledger"),
    ("Less: disbursements during administration", "=SUMIF('Estate Ledger'!B6:B155,\"Disbursement (money out)\",'Estate Ledger'!E6:E155)", "From the Estate Ledger"),
    ("Less: distributions to beneficiaries", "=SUM(Distributions!E6:E45)", "From the Distributions tab"),
    ("REMAINING ESTATE BALANCE", "=B6+B7-B8-B9", "Should match the estate bank account when complete"),
    ("", None, ""),
    ("DEBTS", None, ""),
    ("Total debts identified", "=SUM(Debts!D6:D45)", "From the Debts tab"),
    ("Total debts paid", "=SUM(Debts!F6:F45)", "From the Debts tab"),
    ("Debts remaining", "=B13-B14", ""),
    ("", None, ""),
    ("CLOSING CHECKLIST", None, ""),
    ("All debts and taxes paid", "", "Mark Done when true"),
    ("Final tax returns filed (final 1040, estate 1041)", "", ""),
    ("All distributions made and releases signed", "", ""),
    ("Final accounting shared with beneficiaries / filed with court", "", ""),
    ("Estate bank account closed", "", ""),
]
r = 5
for label, formula, note in lines:
    a = ws.cell(row=r, column=1, value=label)
    a.font = Font(name=SANS, size=11, color=NAVY)
    if label.isupper() and label:
        a.font = Font(name=MONO, bold=True, size=10, color=GOLD if "CHECKLIST" not in label and "DEBTS" not in label else GOLD)
        for col in range(1, 4):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=ZEBRA)
    if label.startswith("REMAINING"):
        a.font = Font(name=SANS, bold=True, size=12, color=NAVY)
        for col in range(1, 4):
            ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=SAGE)
            ws.cell(row=r, column=col).border = NAVY_TOP
    if formula:
        b = ws.cell(row=r, column=2, value=formula)
        b.number_format = MONEY
        b.font = Font(name=SANS, bold=label.startswith("REMAINING"), size=11, color=NAVY)
    n = ws.cell(row=r, column=3, value=note)
    n.font = Font(name=SANS, size=10, italic=True, color=INK_55)
    r += 1
dropdown(ws, STATUS, "B", 18, 22)
ws.cell(row=r + 1, column=1, value=DISCLAIMER).font = Font(name=SANS, size=9, italic=True, color=INK_55)

wb.save("/home/user/hello-world/executor-kit/Executor_Estate_Kit.xlsx")
print("Workbook written.")
